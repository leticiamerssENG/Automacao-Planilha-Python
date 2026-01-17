# Ler dados de planilha
# Automatizar entrada de dados em planilhas
# Inserir dados de qualquer fonte (Word, Banco de Dados, etc)


from openpyxl import load_workbook, Workbook

#Ler dados da planilha
planilha_vendas = load_workbook('vendas_de_lanches.xlsx') #carrega a planilha

pagina_vendas = planilha_vendas['Sheet'] #seleciona a aba 'Sheet'

for linha in pagina_vendas.iter_rows (values_only=True): #iter_rows itera em cada linha
    print(linha)


# Automatizar entrada de dados em planilhas
# Inserir dados de qualquer fonte (Word, Banco de Dados, etc)   

planilha_contas =  Workbook()
pagina1 = planilha_contas.active

with open('anotacoes.txt', 'r', encoding='utf-8') as arquivo:
    for linha in arquivo:
        pagina1.append(linha.split(','))


planilha_contas.save('contas_a_pagar.xlsx')